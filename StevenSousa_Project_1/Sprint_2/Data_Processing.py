"""
Author: Steven Sousa
Prof.: John Santore
Institution: Bridgewater State University - COMP490 - Senior Design & Development
Version: 06Feb2025

This file will parse the json files containing jobs and save them to the DB created in main.py
"""
# Import dependencies
import json
import openpyxl
from typing import Tuple


def process_json(filename: str, existing_job_ids: set, sheet) -> None:
    """
    This function will parse rapid_jobs2.json and save it to a spreadsheet.

    :param filename: rapid_jobs2.json
    :param spreadsheet_name: jobs.xlsx
    :return: jobs.xlsx
    """
    with open(filename) as file:
        for line in file:
            try:
                json_objects = json.loads(line.strip())

                # Ensure json_objects is a list before processing
                if isinstance(json_objects, list):
                    for obj in json_objects:
                        job_id = obj.get('id', 'N/A')
                        job_title = obj.get('title', 'N/A')
                        company_name = obj.get('company', 'N/A')
                        job_description = obj.get('description', 'N/A')
                        location = obj.get('location', 'N/A')
                        min_salary = obj.get('min_salary', 0)
                        max_salary = obj.get('max_salary', 0)
                        salary_time = obj.get('salary_time', 'yearly')
                        posted_date = obj.get('datePosted', 'N/A')

                        # Extract first job provider URL
                        job_providers = obj.get('jobProviders', [])  # Default to empty list if missing
                        url = job_providers[0].get('url', 'URL Not Found') if job_providers else 'URL Not Found'
                        remote = True if 'remote' in location.lower() else False
                        job_tuple = (job_id, job_title, company_name, job_description, location, min_salary,
                                     max_salary, salary_time, posted_date, url, remote)

                        if job_id not in existing_job_ids:
                            sheet.append(job_tuple)
                            existing_job_ids.add(job_id)
                else:
                    print("Unexpected format: Not a list.", json_objects)
            except json.JSONDecodeError as e:
                print("Error decoding JSON:", e)


def load_existing_job_ids(sheet: openpyxl.Workbook) -> set:
    """
    This function will load all existing job ids from the spreadsheet.
    :param sheet: jobs.xlsx
    :return: set of existing job ids
    """
    existing_job_ids = set()
    for row in sheet.iter_rows(min_row=2):
        job_id = row[0].value
        if job_id:
            existing_job_ids.add(job_id)
    return existing_job_ids


def json_to_excel(filename: str, spreadsheet_name='jobs.xlsx') -> None:
    """
    This function will load the existing spreadsheet or create a new one and save the jobs to it.
    :param filename: rapid_jobs2.json
    :param spreadsheet_name: jobs.xlsx
    :return: jobs.xlsx
    """
    try:
        # First, try to load the existing spreadsheet
        workbook = openpyxl.load_workbook(spreadsheet_name)
        sheet = workbook.active
        print(f"Spreadsheet '{spreadsheet_name}' loaded successfully.")
    except FileNotFoundError:
        # If file not found, create one
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Job ID", "Job Title", "Company Name", "Job Description", "Location", "Min Salary",
                      "Max Salary", "Salary Time", "Posted Date", "URL", "Remote"])
        print(f"Spreadsheet '{spreadsheet_name}' created successfully.")

    existing_job_ids = load_existing_job_ids(sheet)
    process_json(filename, existing_job_ids, sheet)

    try:
        workbook.save(spreadsheet_name)
        print(f"Spreadsheet '{spreadsheet_name}' saved successfully.")
    except Exception as e:
        print(f"Error saving spreadsheet '{spreadsheet_name}': {e}")
